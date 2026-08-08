"""Deterministic XLSX/XLS extraction."""

from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
import re
from typing import Any

from ..classification import ContentClassifier
from ..model import Cell, Image, Sheet, SourceLocation, Workbook


_HEADER_WORDS = {
    "#",
    "no",
    "name",
    "description",
    "status",
    "date",
    "code",
    "function",
    "feature",
    "test",
    "issue",
    "owner",
    "deliverable",
    "tool",
    "role",
    "type",
}


def extract_xlsx(
    path: str | Path,
    *,
    classifier: ContentClassifier | None = None,
    sheet_rules: dict[str, Any] | None = None,
    formula_mode: str = "values",
) -> Workbook:
    source = Path(path)
    if source.suffix.casefold() != ".xlsx":
        raise ValueError(f"expected .xlsx input, got {source.suffix or '<none>'}")
    return extract_workbook(source, classifier=classifier, sheet_rules=sheet_rules, formula_mode=formula_mode)


def extract_xls(
    path: str | Path,
    *,
    classifier: ContentClassifier | None = None,
    sheet_rules: dict[str, Any] | None = None,
) -> Workbook:
    source = Path(path)
    if source.suffix.casefold() != ".xls":
        raise ValueError(f"expected .xls input, got {source.suffix or '<none>'}")
    return extract_workbook(source, classifier=classifier, sheet_rules=sheet_rules)


def extract_workbook(
    path: str | Path,
    *,
    classifier: ContentClassifier | None = None,
    sheet_rules: dict[str, Any] | None = None,
    formula_mode: str = "values",
) -> Workbook:
    """Extract a workbook without saving or modifying its source.

    ``formula_mode`` is ``values`` by default: cached/calculated values are
    represented as cell values and formula text is retained in ``Cell.formula``
    when the source format exposes it.  ``formulas`` makes XLSX formula text the
    cell value; legacy XLS files still expose the value returned by xlrd because
    xlrd 2.x does not expose the original formula expression.
    """

    if formula_mode not in {"values", "formulas"}:
        raise ValueError("formula_mode must be 'values' or 'formulas'")
    source = Path(path)
    suffix = source.suffix.casefold()
    if suffix == ".xlsx":
        workbook = _extract_xlsx(source, sheet_rules or {}, formula_mode)
    elif suffix == ".xls":
        workbook = _extract_xls(source, sheet_rules or {})
    else:
        raise ValueError(f"unsupported workbook format: {source.suffix or '<none>'}")
    if classifier is not None:
        apply_classifier(workbook, classifier)
    return workbook


def apply_classifier(workbook: Workbook, classifier: ContentClassifier) -> None:
    for sheet in workbook.sheets:
        for row in sheet.rows:
            for cell in row:
                if not cell.is_empty:
                    classifier.classify_cell(cell)


def _extract_xlsx(source: Path, sheet_rules: dict[str, Any], formula_mode: str) -> Workbook:
    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover - packaging failure
        raise RuntimeError("openpyxl is required for XLSX extraction") from exc

    try:
        value_book = openpyxl.load_workbook(source, data_only=True, read_only=False)
        formula_book = openpyxl.load_workbook(source, data_only=False, read_only=False)
    except Exception as exc:
        raise ValueError(f"could not parse XLSX {source}: {exc}") from exc

    result = Workbook(
        source_path=str(source),
        format="xlsx",
        metadata={"formula_mode": formula_mode, "source_not_modified": True},
    )
    try:
        for value_sheet in value_book.worksheets:
            formula_sheet = formula_book[value_sheet.title]
            rule = _sheet_rule(sheet_rules, value_sheet.title)
            max_row = max(value_sheet.max_row, formula_sheet.max_row)
            max_column = max(value_sheet.max_column, formula_sheet.max_column)
            rows: list[list[Cell]] = []
            for row_number in range(1, max_row + 1):
                row: list[Cell] = []
                for column_number in range(1, max_column + 1):
                    value_cell = value_sheet.cell(row_number, column_number)
                    formula_cell = formula_sheet.cell(row_number, column_number)
                    formula = formula_cell.value if isinstance(formula_cell.value, str) and formula_cell.value.startswith("=") else None
                    value = formula if formula_mode == "formulas" and formula is not None else _json_cell_value(value_cell.value)
                    row.append(
                        Cell(
                            value=value,
                            row=row_number,
                            column=column_number,
                            source_location=SourceLocation(
                                kind="xlsx_cell",
                                sheet_name=value_sheet.title,
                                row=row_number,
                                column=column_number,
                            ),
                            original_text=_display_value(value),
                            formula=formula,
                            metadata={"formula_mode": formula_mode},
                        )
                    )
                rows.append(row)
            merged = [str(region) for region in value_sheet.merged_cells.ranges]
            sheet = Sheet(
                name=value_sheet.title,
                rows=rows,
                merged_regions=merged,
                detected_header_rows=_header_rows(rows, rule),
                source_location=SourceLocation(kind="xlsx_sheet", sheet_name=value_sheet.title),
                metadata={"max_row": max_row, "max_column": max_column},
            )
            sheet.images.extend(_xlsx_images(value_sheet))
            result.sheets.append(sheet)
    finally:
        value_book.close()
        formula_book.close()
    return result


def _extract_xls(source: Path, sheet_rules: dict[str, Any]) -> Workbook:
    try:
        import xlrd
    except ImportError as exc:  # pragma: no cover - packaging failure
        raise RuntimeError("xlrd is required for legacy XLS extraction") from exc

    try:
        book = xlrd.open_workbook(source, formatting_info=True, on_demand=False)
    except Exception as exc:
        raise ValueError(f"could not parse XLS {source}: {exc}") from exc

    result = Workbook(
        source_path=str(source),
        format="xls",
        metadata={
            "formula_mode": "values",
            "source_not_modified": True,
            "formula_note": "xlrd exposes legacy XLS formula results, not original expressions",
        },
    )
    for source_sheet in book.sheets():
        rule = _sheet_rule(sheet_rules, source_sheet.name)
        rows: list[list[Cell]] = []
        for row_number in range(1, source_sheet.nrows + 1):
            row: list[Cell] = []
            for column_number in range(1, source_sheet.ncols + 1):
                source_cell = source_sheet.cell(row_number - 1, column_number - 1)
                value = _xlrd_value(source_cell, book.datemode, xlrd)
                row.append(
                    Cell(
                        value=value,
                        row=row_number,
                        column=column_number,
                        source_location=SourceLocation(
                            kind="xls_cell",
                            sheet_name=source_sheet.name,
                            row=row_number,
                            column=column_number,
                        ),
                        original_text=_display_value(value),
                        metadata={"formula_mode": "values", "xlrd_cell_type": source_cell.ctype},
                    )
                )
            rows.append(row)
        merged = [_xlrd_region(region) for region in source_sheet.merged_cells]
        result.sheets.append(
            Sheet(
                name=source_sheet.name,
                rows=rows,
                merged_regions=merged,
                detected_header_rows=_header_rows(rows, rule),
                source_location=SourceLocation(kind="xls_sheet", sheet_name=source_sheet.name),
                metadata={"nrows": source_sheet.nrows, "ncols": source_sheet.ncols},
            )
        )
    return result


def _sheet_rule(sheet_rules: dict[str, Any], name: str) -> dict[str, Any]:
    if not isinstance(sheet_rules, dict):
        return {}
    value = sheet_rules.get(name)
    if isinstance(value, dict):
        return value
    for pattern_rule in sheet_rules.get("__patterns__", []) or []:
        if not isinstance(pattern_rule, dict) or not pattern_rule.get("pattern"):
            continue
        try:
            if re.search(str(pattern_rule["pattern"]), name, re.IGNORECASE):
                return pattern_rule
        except re.error:
            continue
    return {}


def _header_rows(rows: list[list[Cell]], rule: dict[str, Any]) -> list[int]:
    configured = rule.get("header_rows")
    if isinstance(configured, list) and all(isinstance(value, int) for value in configured):
        return list(configured)
    candidates: list[tuple[int, int]] = []
    for row_number, row in enumerate(rows[:30], start=1):
        values = [str(cell.value).strip() for cell in row if not cell.is_empty]
        if len(values) < 2:
            continue
        words = set()
        for value in values:
            words.update(re.findall(r"[a-z0-9#]+", value.casefold()))
        label_count = sum(_looks_like_label(value) for value in values)
        keyword_count = len(words & _HEADER_WORDS)
        if keyword_count >= 1 and label_count >= 2:
            # Prefer the row with the richest set of short labels.  Metadata
            # rows such as Project Name/Project Code appear before the actual
            # table header in the supplied templates.
            candidates.append((label_count + keyword_count, row_number))
    if not candidates:
        return []
    return [max(candidates, key=lambda item: (item[0], item[1]))[1]]


def _looks_like_label(value: str) -> bool:
    return bool(value) and len(value) <= 80 and not value.startswith("=")


def _xlsx_images(sheet: Any) -> list[Image]:
    images: list[Image] = []
    for index, image in enumerate(getattr(sheet, "_images", []), start=1):
        anchor = getattr(image, "anchor", None)
        row = getattr(getattr(anchor, "_from", None), "row", None)
        column = getattr(getattr(anchor, "_from", None), "col", None)
        images.append(
            Image(
                relationship_id=f"image{index}",
                description=getattr(image, "path", None),
                source_location=SourceLocation(
                    kind="xlsx_image",
                    sheet_name=sheet.title,
                    row=(row + 1) if row is not None else None,
                    column=(column + 1) if column is not None else None,
                ),
                filename=getattr(image, "path", None),
                metadata={"embedded": True},
            )
        )
    return images


def _xlrd_region(region: tuple[int, int, int, int]) -> str:
    row_start, row_end, column_start, column_end = region
    return f"{_column_name(column_start + 1)}{row_start + 1}:{_column_name(column_end)}{row_end}"


def _xlrd_value(cell: Any, datemode: int, xlrd: Any) -> Any:
    if cell.ctype in (xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK):
        return None
    if cell.ctype == xlrd.XL_CELL_DATE:
        try:
            value = xlrd.xldate_as_datetime(cell.value, datemode)
            return value.isoformat()
        except Exception:
            return cell.value
    if cell.ctype == xlrd.XL_CELL_BOOLEAN:
        return bool(cell.value)
    if cell.ctype == xlrd.XL_CELL_ERROR:
        return f"#ERROR:{cell.value}"
    return cell.value


def _json_cell_value(value: Any) -> Any:
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return value


def _display_value(value: Any) -> str:
    return "" if value is None else str(value)


def _column_name(number: int) -> str:
    result = ""
    while number > 0:
        number, remainder = divmod(number - 1, 26)
        result = chr(65 + remainder) + result
    return result or "?"
